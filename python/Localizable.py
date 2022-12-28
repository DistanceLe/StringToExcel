# -*- coding:utf-8 -*-

import os
from optparse import OptionParser
from LocalizableStringsFileUtil import LocalizableStringsFileUtil
#import pyExcelerator
from openpyxl import Workbook

#Add command option
def addParser():
    parser = OptionParser()

    parser.add_option("-f", "--filesDirectory",
                      help="Localizable.strings files directory.",
                      metavar="filesDirectory")

    parser.add_option("-t", "--targetFilePath",
                      help="Target File (xls) Path.",
                      metavar="targetFilePath")

    (options, args) = parser.parse_args()

    return options


# Start convert Localizable.strings to xls
def startConvert(options):
    directory = options.filesDirectory
    targetFilePath = options.targetFilePath

    if directory is not None:
        index = 0
        if targetFilePath is not None:
            

            for parent, dirnames, filenames in os.walk(directory):
                index = 0

                for dirname in dirnames:
                    # KeyName & CountryCode
                    #workbook = pyExcelerator.Workbook() #生成一个工作簿
                    wb = Workbook()

                    #ws = workbook.add_sheet('Localizable.strings') #加入一个工作表（sheet）
                    ws = wb.active;

                    if index == 0:
                        #ws.write(0,0,'keyName')
                        ws['A1'] = 'keyName'
                    conturyCode = dirname.split('.')[0] #得到国家的名字
                    #ws.write(0,index+1,conturyCode) #在第一行 第index+1列 写入国家名字
                    ws['B1'] = conturyCode



                    # Key & Value
                    path = directory+'/'+dirname+'/Localizable.strings'
                    (keys, values) = LocalizableStringsFileUtil.getKeysAndValues(path)
                    for x in range(len(keys)):
                        key = keys[x]
                        value = values[x]
                        #if (index == 0):
                        #ws.write(x+1, 0, key)
                        #ws.write(x+1, 1, value)

                        AIndex = 'A%d' % (x+2)
                        ws[AIndex] = key
                        BIndex = "B%d" % (x+2)
                        ws[BIndex] = value
                        #print "index =====  %s" % (x+2)+value
                        #filePath = targetFilePath + "/Localizable_" + conturyCode + ".xlsx"
                        #wb.save(filePath)
                        #ws.append([key, value])
                        #ws["A3"] = key 
                        #ws.append([1, 2, 3])
                        #ws['A1'] = 42
                        #ws['A2'] = datetime.datetime.now()

                        #ws['B'+(x+1)] = value

                        #else:
                         #   ws.write(x+1, index + 1, value)
                    #index += 1
                    filePath = targetFilePath + "/Localizable_" + conturyCode + ".xlsx"
                    wb.save(filePath)  #保存Excel 表

                    print "Convert successfully! you can see xls file in %s" % (filePath)

            

        else:
            print "Target file path can not be empty! try -h for help."
    else:
        print "Localizable.strings files directory can not be empty! try -h for help."


def main():
    options = addParser()
    startConvert(options)

main()